"""
Prime Rush Release Tester
=========================
Usage:
    python tester.py
        -- looks for "Content of Prime Rush (BRX).xlsx" in your Downloads folder

    python tester.py "path/to/Content of Prime Rush (BRX).xlsx"
        -- use a specific file

Results are saved as "test-results-YYYY-MM-DD-HHMM.xlsx" in the same folder
you run the script from.
"""

import sys
import os
import json
import webbrowser
import threading
from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify, request, send_from_directory
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

app = Flask(__name__, static_folder=str(Path(__file__).parent))

ITEMS_DATA = {}


# ----------------------------------------------------------------
# Load + parse the content xlsx
# ----------------------------------------------------------------
def load_items(xlsx_path: str) -> tuple[dict, int]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    sheet_name = 'Main' if 'Main' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.values)
    wb.close()

    if not rows:
        raise ValueError('Sheet is empty.')

    # Auto-detect header row (search first 5 rows)
    header_row_idx = 0
    for i, row in enumerate(rows[:5]):
        if any(str(c).strip().lower() == 'type' for c in row if c is not None):
            header_row_idx = i
            break

    headers = [str(c).strip() if c is not None else '' for c in rows[header_row_idx]]

    def col(name: str) -> int:
        for i, h in enumerate(headers):
            if h.lower() == name.lower():
                return i
        visible = [h for h in headers if h]
        raise ValueError(f'Column "{name}" not found. Headers found: {visible}')

    type_idx        = col('Type')
    name_idx        = col('Name')
    rarity_idx      = col('Rarity')
    skin_id_idx     = col('Skin ID')
    status_idx      = col('Status')
    visibility_idx  = col('Prime Rush Visibility')
    acquisition_idx = col('Prime Rush Current Acquisitions')

    def get(row, idx: int) -> str:
        if idx < len(row) and row[idx] is not None:
            return str(row[idx]).strip()
        return ''

    grouped: dict[str, list] = {}

    for row in rows[header_row_idx + 1:]:
        if not any(c for c in row if c is not None):
            continue  # blank row

        status     = get(row, status_idx).lower()
        visibility = get(row, visibility_idx).lower()

        is_active  = status in ('on going', 'ongoing')
        is_visible = visibility == 'visible'

        if not is_active or not is_visible:
            continue

        item_type = get(row, type_idx) or 'Unknown'
        name      = get(row, name_idx)

        if not name:
            continue

        item = {
            'type':        item_type,
            'name':        name,
            'rarity':      get(row, rarity_idx),
            'skinId':      get(row, skin_id_idx),
            'acquisition': get(row, acquisition_idx),
        }

        grouped.setdefault(item_type, []).append(item)

    total = sum(len(v) for v in grouped.values())
    return grouped, total


# ----------------------------------------------------------------
# Routes
# ----------------------------------------------------------------
@app.route('/')
def index():
    return send_from_directory(str(Path(__file__).parent), 'index.html')


@app.route('/items')
def get_items():
    return jsonify(ITEMS_DATA)


@app.route('/save', methods=['POST'])
def save_results():
    try:
        payload  = request.get_json()
        results  = payload['results']
        summary  = payload['summary']

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Test Results'

        # ---- Summary block ----
        ws.append(['PRIME RUSH RELEASE TEST SUMMARY'])
        ws.append(['Date',        datetime.now().strftime('%m/%d/%Y %H:%M')])
        ws.append(['Total Items', summary['total']])
        ws.append(['Passed',      summary['passed']])
        ws.append(['Bugs Found',  summary['bugs']])
        ws.append(['Skipped',     summary['skipped']])
        ws.append([])  # spacer row

        # ---- Table header ----
        ws.append(['Category', 'Item Name', 'Rarity', 'Acquisition', 'Result', 'Bug Notes'])
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font       = Font(bold=True, color='FFFFFF')
            cell.fill       = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
            cell.alignment  = Alignment(horizontal='center')

        # ---- Item rows ----
        FILL = {
            'PASS': PatternFill(start_color='b7e1cd', end_color='b7e1cd', fill_type='solid'),
            'BUG':  PatternFill(start_color='f4c7c3', end_color='f4c7c3', fill_type='solid'),
            'SKIP': PatternFill(start_color='fce8b2', end_color='fce8b2', fill_type='solid'),
        }

        for r in results:
            ws.append([
                r['type'],
                r['name'],
                r['rarity'],
                r['acquisition'],
                r['status'],
                r.get('bugNotes', ''),
            ])
            result_cell      = ws.cell(row=ws.max_row, column=5)
            result_cell.fill = FILL.get(r['status'], FILL['SKIP'])

        # ---- Title styling ----
        ws['A1'].font = Font(bold=True, size=13)

        # ---- Column widths ----
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=8
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 55)

        filename    = f'test-results-{datetime.now().strftime("%Y-%m-%d-%H%M")}.xlsx'
        output_path = Path(filename)
        wb.save(output_path)

        return jsonify({'success': True, 'filename': filename})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ----------------------------------------------------------------
# Entry point
# ----------------------------------------------------------------
def find_content_file() -> str | None:
    candidates = [
        Path.home() / 'Downloads' / 'Content of Prime Rush (BRX).xlsx',
        Path('Content of Prime Rush (BRX).xlsx'),
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    return None


def open_browser_after_delay():
    import time
    time.sleep(1.2)
    webbrowser.open('http://localhost:5000')


if __name__ == '__main__':
    # Resolve xlsx path
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    else:
        xlsx_path = find_content_file()
        if not xlsx_path:
            print('Could not find the content file automatically.')
            print('Usage: python tester.py "Content of Prime Rush (BRX).xlsx"')
            sys.exit(1)

    print(f'Loading: {xlsx_path}')

    try:
        grouped, total = load_items(xlsx_path)
    except Exception as e:
        print(f'Error reading file: {e}')
        sys.exit(1)

    ITEMS_DATA.update({
        'success':    True,
        'data':       grouped,
        'totalItems': total,
    })

    print(f'Loaded {total} items across {len(grouped)} categories.')
    print('Opening http://localhost:5000 ...')
    print('Press Ctrl+C to stop the server when done.\n')

    threading.Thread(target=open_browser_after_delay, daemon=True).start()
    app.run(host='127.0.0.1', port=5000, debug=False)
